#!/usr/bin/env python3
"""
Import ComplianceForge Secure Controls Framework (SCF)
Downloads and imports the complete SCF control library
"""

import json
import logging
import psycopg2
import requests
import pandas as pd
from typing import Dict, List, Any, Optional
from pathlib import Path
import yaml
import openpyxl

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class SCFImporter:
    """Import ComplianceForge Secure Controls Framework"""

    def __init__(self, db_config: Dict[str, str]):
        self.db_config = db_config
        self.conn = psycopg2.connect(**db_config)

    def download_scf_excel(self, output_path: Path) -> Path:
        """
        Download SCF Excel file

        Note: ComplianceForge requires registration for download.
        You'll need to manually download from:
        https://www.complianceforge.com/secure-controls-framework-scf/

        Or use their API if you have credentials.
        """
        logger.warning("SCF download requires manual steps:")
        logger.warning("1. Visit https://www.complianceforge.com/secure-controls-framework-scf/")
        logger.warning("2. Download SCF_2024.1.xlsx")
        logger.warning(f"3. Place it at: {output_path}")

        if not output_path.exists():
            raise FileNotFoundError(f"Please download SCF Excel file to {output_path}")

        return output_path

    def create_framework(self) -> str:
        """Create SCF framework entry"""
        with self.conn.cursor() as cur:
            cur.execute("""
                INSERT INTO compliance_frameworks (name, version, source, description, framework_url)
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT (name) DO UPDATE SET
                    version = EXCLUDED.version,
                    description = EXCLUDED.description
                RETURNING id
            """, (
                "ComplianceForge-SCF",
                "2024.1",
                "ComplianceForge",
                "Secure Controls Framework - Unified security & privacy controls",
                "https://www.complianceforge.com/scf"
            ))

            framework_id = cur.fetchone()[0]
            self.conn.commit()

        return framework_id

    def create_domain(self, framework_id: str, domain_code: str, domain_name: str) -> str:
        """Create or get domain ID"""
        with self.conn.cursor() as cur:
            cur.execute("""
                INSERT INTO control_domains (framework_id, domain_code, domain_name)
                VALUES (%s, %s, %s)
                ON CONFLICT (framework_id, domain_code) DO UPDATE SET
                    domain_name = EXCLUDED.domain_name
                RETURNING id
            """, (framework_id, domain_code, domain_name))

            domain_id = cur.fetchone()[0]
            self.conn.commit()

        return domain_id

    def parse_control_type(self, scf_type: str) -> str:
        """Map SCF control type to our taxonomy"""
        scf_type_lower = (scf_type or "").lower()

        if "preventive" in scf_type_lower or "prevent" in scf_type_lower:
            return "preventive"
        elif "detective" in scf_type_lower or "detect" in scf_type_lower:
            return "detective"
        elif "corrective" in scf_type_lower or "correct" in scf_type_lower:
            return "corrective"
        elif "directive" in scf_type_lower:
            return "directive"
        else:
            return "preventive"  # Default

    def parse_mappings(self, mapping_str: str) -> Dict[str, List[str]]:
        """
        Parse SCF mappings to other frameworks

        Example mapping_str:
        "NIST 800-53: AC-1, AC-2 | ISO 27001: A.9.1.1 | PCI DSS: 8.1"
        """
        if not mapping_str:
            return {}

        mappings = {}

        # Split by |
        parts = mapping_str.split('|')

        for part in parts:
            if ':' not in part:
                continue

            framework, controls = part.split(':', 1)
            framework = framework.strip()
            control_list = [c.strip() for c in controls.split(',')]

            mappings[framework] = control_list

        return mappings

    def import_scf_from_excel(self, excel_path: Path):
        """Import SCF controls from Excel file"""
        logger.info(f"Loading SCF Excel file: {excel_path}")

        # Load Excel
        workbook = openpyxl.load_workbook(excel_path)

        # SCF has controls in the "Controls" sheet
        if "Controls" not in workbook.sheetnames:
            logger.error(f"No 'Controls' sheet found. Available sheets: {workbook.sheetnames}")
            raise ValueError("Invalid SCF Excel format")

        sheet = workbook["Controls"]

        # Create framework
        framework_id = self.create_framework()

        # Parse header row to find column indices
        header_row = [cell.value for cell in sheet[1]]
        logger.info(f"Excel columns: {header_row}")

        # Expected columns (may vary by SCF version)
        col_mapping = {
            'domain': None,
            'control_id': None,
            'control_title': None,
            'control_specification': None,
            'control_type': None,
            'nist_mapping': None,
            'iso_mapping': None,
        }

        # Find column indices (case-insensitive partial match)
        for idx, header in enumerate(header_row):
            header_lower = (header or "").lower()

            if 'domain' in header_lower and 'name' not in header_lower:
                col_mapping['domain'] = idx
            elif 'control id' in header_lower or 'scf id' in header_lower:
                col_mapping['control_id'] = idx
            elif 'control title' in header_lower or 'name' in header_lower:
                col_mapping['control_title'] = idx
            elif 'specification' in header_lower or 'description' in header_lower:
                col_mapping['control_specification'] = idx
            elif 'control type' in header_lower or 'type' in header_lower:
                col_mapping['control_type'] = idx
            elif 'nist' in header_lower and 'mapping' in header_lower:
                col_mapping['nist_mapping'] = idx
            elif 'iso' in header_lower and 'mapping' in header_lower:
                col_mapping['iso_mapping'] = idx

        logger.info(f"Column mapping: {col_mapping}")

        # Check if we found required columns
        if not all([col_mapping['domain'], col_mapping['control_id'], col_mapping['control_title']]):
            raise ValueError("Could not find required columns in Excel file")

        # Process rows
        domains_cache = {}
        control_count = 0

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                domain_code = row[col_mapping['domain']]
                control_id = row[col_mapping['control_id']]
                control_title = row[col_mapping['control_title']]

                if not domain_code or not control_id:
                    continue  # Skip empty rows

                # Create domain if not exists
                if domain_code not in domains_cache:
                    domains_cache[domain_code] = self.create_domain(
                        framework_id,
                        domain_code,
                        f"SCF Domain {domain_code}"
                    )

                domain_id = domains_cache[domain_code]

                # Get control specification
                control_spec = row[col_mapping['control_specification']] if col_mapping['control_specification'] else ""

                # Get control type
                control_type_raw = row[col_mapping['control_type']] if col_mapping['control_type'] else "preventive"
                control_type = self.parse_control_type(control_type_raw)

                # Get mappings
                nist_mapping = row[col_mapping['nist_mapping']] if col_mapping['nist_mapping'] else ""
                iso_mapping = row[col_mapping['iso_mapping']] if col_mapping['iso_mapping'] else ""

                metadata = {
                    'scf_version': '2024.1',
                    'nist_mapping': nist_mapping,
                    'iso_mapping': iso_mapping,
                    'control_type_raw': control_type_raw
                }

                # Insert control
                with self.conn.cursor() as cur:
                    cur.execute("""
                        INSERT INTO controls (
                            domain_id,
                            control_code,
                            control_name,
                            control_description,
                            control_type,
                            metadata
                        ) VALUES (%s, %s, %s, %s, %s, %s)
                        ON CONFLICT (domain_id, control_code) DO UPDATE SET
                            control_name = EXCLUDED.control_name,
                            control_description = EXCLUDED.control_description,
                            control_type = EXCLUDED.control_type,
                            metadata = EXCLUDED.metadata
                    """, (
                        domain_id,
                        control_id,
                        control_title,
                        control_spec or "",
                        control_type,
                        json.dumps(metadata)
                    ))

                control_count += 1

                if control_count % 50 == 0:
                    logger.info(f"Imported {control_count} SCF controls...")
                    self.conn.commit()

            except Exception as e:
                logger.error(f"Error processing row {row_idx}: {e}")
                continue

        self.conn.commit()

        logger.info(f"✅ SCF import complete! Imported {control_count} controls")
        return control_count


def main():
    import argparse

    parser = argparse.ArgumentParser(description='Import ComplianceForge SCF')
    parser.add_argument('--config', required=True, help='Database config YAML file')
    parser.add_argument('--scf-excel', required=True, help='Path to SCF Excel file (SCF_2024.1.xlsx)')

    args = parser.parse_args()

    # Load config
    with open(args.config, 'r') as f:
        config = yaml.safe_load(f)

    # Import
    importer = SCFImporter(config['database'])
    total = importer.import_scf_from_excel(Path(args.scf_excel))

    print(f"\n✅ Successfully imported {total} ComplianceForge SCF controls!")


if __name__ == "__main__":
    main()
